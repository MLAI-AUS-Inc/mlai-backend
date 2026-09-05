from __future__ import annotations

import base64
from datetime import datetime, timezone
from types import SimpleNamespace
from unittest.mock import patch
from urllib.parse import urlparse

import requests
from django.test import SimpleTestCase, override_settings
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase

from integrations.services.reconciliation import (
    ReconciliationReportService,
    StripeAPIError,
    _dollars,
)
from roo.models import PointsAdmin


def _sum_split_amounts(brief: str) -> float:
    """Sum the Amount column of every split-table data row across the brief."""
    total = 0.0
    in_split = False
    for line in brief.splitlines():
        if line.startswith("| What (account) |"):
            in_split = True
            continue
        if not in_split:
            continue
        if line.startswith("|---") or line.startswith("| ---"):
            continue
        if "TOTAL" in line:
            in_split = False
            continue
        if line.startswith("|"):
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            if len(cells) >= 4:
                amt = cells[3].replace("**", "").replace(",", "").replace("$", "")
                try:
                    total += float(amt)
                except ValueError:
                    pass
    return round(total, 2)


class FakeResponse:
    def __init__(self, payload, status_code=200):
        self.payload = payload
        self.status_code = status_code

    def json(self):
        return self.payload

    def raise_for_status(self):
        if self.status_code >= 400:
            raise requests.HTTPError(f"HTTP {self.status_code}")


class FakeSession:
    def __init__(self, handler):
        self.handler = handler
        self.calls = []

    def get(self, url, headers, params, timeout):
        call = {"path": urlparse(url).path, "headers": dict(headers), "params": dict(params)}
        self.calls.append(call)
        result = self.handler(call["path"], call["params"])
        return result if isinstance(result, FakeResponse) else FakeResponse(result)


def _charge(txn_id, amount, fee, event_api_id, email, description, created=1_780_000_000):
    return {
        "id": txn_id,
        "type": "charge",
        "amount": amount,
        "fee": fee,
        "net": amount - fee,
        "currency": "aud",
        "created": created,
        "source": {
            "id": "ch_" + txn_id,
            "description": description,
            "created": created,
            "metadata": {"event_api_id": event_api_id, "email": email} if event_api_id else {"email": email},
            "billing_details": {"email": email},
        },
    }


class ReconciliationServiceTests(SimpleTestCase):
    def _service(self, handler):
        return ReconciliationReportService(
            stripe_api_key="rk_test", base_url="https://stripe.test", session=FakeSession(handler)
        )

    def _window(self):
        return (
            datetime(2026, 6, 1, tzinfo=timezone.utc),
            datetime(2026, 7, 1, tzinfo=timezone.utc),
        )

    def test_payout_driven_report_ties_out_and_groups_by_event(self):
        payouts = {
            "po_A": {"id": "po_A", "amount": 6725, "currency": "aud", "arrival_date": 1_780_600_000, "status": "paid"},
        }
        txns = {
            "po_A": [
                _charge("bt1", 5000, 175, "evt-1", "a@example.com", "MLAI Workshop"),
                _charge("bt2", 2000, 100, "evt-2", "b@example.com", "MLAI Mixer"),
                {"id": "bt_po", "type": "payout", "amount": -6725, "net": -6725, "currency": "aud"},
            ]
        }

        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": list(payouts.values()), "has_more": False}
            if path == "/v1/balance_transactions":
                return {"data": txns[params["payout"]], "has_more": False}
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=True)

        self.assertEqual(report["payout_count"], 1)
        self.assertEqual(report["charge_count"], 2)
        self.assertEqual(report["unmatched_charge_count"], 0)

        p = report["payouts"][0]
        self.assertEqual(p["payout_id"], "po_A")
        self.assertEqual(p["gross_cents"], 7000)
        self.assertEqual(p["stripe_fee_cents"], 275)
        self.assertEqual(p["charge_net_cents"], 6725)
        self.assertEqual(p["deposit_cents"], 6725)
        self.assertEqual(p["deposit_cents"], p["payout_amount_cents"])  # tie-out
        self.assertEqual([e["event_name"] for e in p["events"]], ["MLAI Mixer", "MLAI Workshop"])
        self.assertEqual(report["currency_totals"]["AUD"]["deposit"], 67.25)
        self.assertEqual(p["warnings"], [])  # clean payout, no warnings

        brief = base64.b64decode(report["brief"]["content_base64"]).decode("utf-8")
        self.assertIn("po_A", brief)
        self.assertIn("MLAI Workshop", brief)
        self.assertIn("67.25", brief)

        # workbook present and is a real xlsx (zip magic bytes)
        self.assertIn("workbook", report)
        self.assertEqual(base64.b64decode(report["workbook"]["content_base64"])[:2], b"PK")

    def test_unmatched_charge_and_refund_are_surfaced(self):
        def handler(path, params):
            if path == "/v1/payouts":
                return {
                    "data": [{"id": "po_B", "amount": 2400, "currency": "aud", "arrival_date": 1_780_600_000, "status": "paid"}],
                    "has_more": False,
                }
            if path == "/v1/balance_transactions":
                return {
                    "data": [
                        _charge("bt3", 3000, 100, "", "c@example.com", "Mystery ticket"),
                        {"id": "bt_ref", "type": "refund", "amount": -500, "net": -500, "currency": "aud", "description": "refund"},
                        {"id": "bt_po2", "type": "payout", "amount": -2400, "net": -2400, "currency": "aud"},
                    ],
                    "has_more": False,
                }
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=False)

        self.assertEqual(report["unmatched_charge_count"], 1)
        p = report["payouts"][0]
        # charge nets (2900) - refund (500) = deposit (2400) = bank line
        self.assertEqual(p["charge_net_cents"], 2900)
        self.assertEqual(p["refund_net_cents"], -500)
        self.assertEqual(p["deposit_cents"], 2400)
        self.assertEqual(p["deposit_cents"], p["payout_amount_cents"])
        self.assertFalse(any("Tie-out mismatch" in w for w in p["warnings"]))  # ties out despite refund
        self.assertEqual(len(p["refunds"]), 1)
        self.assertTrue(any("no Luma event_api_id" in w for w in p["warnings"]))
        self.assertTrue(any("refunds/adjustments" in w for w in p["warnings"]))
        self.assertNotIn("workbook", report)

    def test_payment_intent_metadata_is_followed_for_luma_event_attribution(self):
        def handler(path, params):
            if path == "/v1/payouts":
                return {
                    "data": [{
                        "id": "po_intent",
                        "amount": 9700,
                        "currency": "aud",
                        "arrival_date": 1_780_600_000,
                        "status": "paid",
                    }],
                    "has_more": False,
                }
            if path == "/v1/balance_transactions":
                return {
                    "data": [{
                        "id": "bt_intent",
                        "type": "charge",
                        "amount": 10000,
                        "fee": 300,
                        "net": 9700,
                        "currency": "aud",
                        "source": {
                            "id": "ch_intent",
                            "payment_intent": "pi_intent",
                            "metadata": {},
                        },
                    }],
                    "has_more": False,
                }
            if path == "/v1/payment_intents/pi_intent":
                return {
                    "id": "pi_intent",
                    "description": "Intent Event",
                    "metadata": {"event_api_id": "evt_intent"},
                }
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(
            since=since,
            until=until,
            include_workbook=False,
        )

        group = report["payouts"][0]["revenue_groups"][0]
        self.assertEqual(group["source_type"], "luma_event")
        self.assertEqual(group["source_id"], "evt_intent")
        self.assertEqual(group["stripe_payment_intent_ids"], ["pi_intent"])

    def test_metadata_poor_charge_uses_unique_luma_captured_order(self):
        class FakeLuma:
            def list_all_events(self):
                return [{"id": "evt_fallback", "name": "Fallback Night"}]

            def get_guest(self, *, event_id, identifier):
                assert event_id == "evt_fallback"
                assert identifier == "guest@example.com"
                return {
                    "event_ticket_orders": [{
                        "id": "order_fallback",
                        "amount": 10000,
                        "currency": "aud",
                        "is_captured": True,
                        "amount_refunded": 0,
                    }]
                }

        def handler(path, params):
            if path == "/v1/payouts":
                return {
                    "data": [{
                        "id": "po_fallback",
                        "amount": 9700,
                        "currency": "aud",
                        "arrival_date": 1_780_600_000,
                        "status": "paid",
                    }],
                    "has_more": False,
                }
            if path == "/v1/balance_transactions":
                return {
                    "data": [_charge(
                        "bt_fallback",
                        10000,
                        300,
                        "",
                        "guest@example.com",
                        "Fallback Night",
                    )],
                    "has_more": False,
                }
            raise AssertionError(path)

        since, until = self._window()
        report = ReconciliationReportService(
            stripe_api_key="rk_test",
            base_url="https://stripe.test",
            session=FakeSession(handler),
            luma_service=FakeLuma(),
        ).build_report(since=since, until=until, include_workbook=False)

        group = report["payouts"][0]["revenue_groups"][0]
        self.assertEqual(group["source_type"], "luma_event")
        self.assertEqual(group["source_id"], "evt_fallback")
        self.assertEqual(group["luma_order_ids"], ["order_fallback"])
        self.assertEqual(
            group["luma_match_methods"],
            ["luma_event_name_email_captured_order_amount"],
        )
        self.assertEqual(report["unmatched_charge_count"], 0)

    def test_ambiguous_luma_orders_remain_unattributed(self):
        class AmbiguousLuma:
            def list_all_events(self):
                return [{"id": "evt_ambiguous", "name": "Ambiguous Night"}]

            def get_guest(self, *, event_id, identifier):
                return {
                    "event_ticket_orders": [
                        {"id": "order_1", "amount": 10000, "currency": "aud", "is_captured": True},
                        {"id": "order_2", "amount": 10000, "currency": "aud", "is_captured": True},
                    ]
                }

        def handler(path, params):
            if path == "/v1/payouts":
                return {
                    "data": [{
                        "id": "po_ambiguous_luma",
                        "amount": 9700,
                        "currency": "aud",
                        "arrival_date": 1_780_600_000,
                        "status": "paid",
                    }],
                    "has_more": False,
                }
            if path == "/v1/balance_transactions":
                return {
                    "data": [_charge(
                        "bt_ambiguous_luma",
                        10000,
                        300,
                        "",
                        "guest@example.com",
                        "Ambiguous Night",
                    )],
                    "has_more": False,
                }
            raise AssertionError(path)

        since, until = self._window()
        report = ReconciliationReportService(
            stripe_api_key="rk_test",
            base_url="https://stripe.test",
            session=FakeSession(handler),
            luma_service=AmbiguousLuma(),
        ).build_report(since=since, until=until, include_workbook=False)

        self.assertEqual(report["unmatched_charge_count"], 1)
        self.assertEqual(
            report["payouts"][0]["revenue_groups"][0]["source_type"],
            "unattributed",
        )

    def test_tie_out_mismatch_warns(self):
        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [{"id": "po_C", "amount": 9999, "currency": "aud", "arrival_date": 1, "status": "paid"}], "has_more": False}
            if path == "/v1/balance_transactions":
                return {"data": [_charge("btx", 5000, 175, "evt-9", "d@example.com", "Ev")], "has_more": False}
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=False)
        self.assertTrue(any("Tie-out mismatch" in w for w in report["payouts"][0]["warnings"]))

    def test_pagination_of_payouts_and_transactions(self):
        # po_1 is returned across two payout pages' worth of balance txns; po_2 on
        # the second payouts page. Exercises both starting_after cursors.
        def handler(path, params):
            if path == "/v1/payouts":
                if params.get("starting_after") == "po_1":
                    return {"data": [{"id": "po_2", "amount": 1900, "currency": "aud", "arrival_date": 2, "status": "paid"}], "has_more": False}
                return {"data": [{"id": "po_1", "amount": 5750, "currency": "aud", "arrival_date": 1, "status": "paid"}], "has_more": True}
            if path == "/v1/balance_transactions":
                po = params["payout"]
                if po == "po_1" and not params.get("starting_after"):
                    return {"data": [_charge("bt_b", 3000, 100, "evt-1", "a@x.com", "A")], "has_more": True}
                if po == "po_1":
                    return {"data": [_charge("bt_d", 3000, 75, "evt-1", "b@x.com", "A")], "has_more": False}
                return {"data": [_charge("bt_c", 2000, 100, "evt-2", "c@x.com", "B")], "has_more": False}
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=False)
        self.assertEqual(report["payout_count"], 2)
        self.assertEqual(report["charge_count"], 3)  # po_1 has 2 (paginated), po_2 has 1

    def test_stripe_401_raises_api_error(self):
        def handler(path, params):
            return FakeResponse({"error": {"message": "bad key"}}, status_code=401)

        since, until = self._window()
        with self.assertRaises(StripeAPIError) as ctx:
            self._service(handler).build_report(since=since, until=until)
        self.assertEqual(ctx.exception.status_code, 401)

    def test_sends_bearer_auth_and_expand(self):
        seen = {}

        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [{"id": "po_z", "amount": 4825, "currency": "aud", "arrival_date": 1, "status": "paid"}], "has_more": False}
            if path == "/v1/balance_transactions":
                seen["bt_params"] = params
                return {"data": [_charge("bt", 5000, 175, "evt-1", "a@x.com", "A")], "has_more": False}
            raise AssertionError(path)

        session = FakeSession(handler)
        service = ReconciliationReportService(stripe_api_key="rk_secret", base_url="https://stripe.test", session=session)
        since, until = self._window()
        service.build_report(since=since, until=until, include_workbook=False)

        self.assertEqual(session.calls[0]["headers"]["Authorization"], "Bearer rk_secret")
        self.assertIn("Stripe-Version", session.calls[0]["headers"])
        self.assertEqual(seen["bt_params"]["expand[]"], "data.source")
        self.assertEqual(seen["bt_params"]["payout"], "po_z")

    def test_brief_split_lines_sum_to_deposit(self):
        # two events + a refund: split lines (gross x2, -fee, -refund) must total the deposit
        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [{"id": "po_S", "amount": 6225, "currency": "aud", "arrival_date": 1_780_600_000, "status": "paid"}], "has_more": False}
            if path == "/v1/balance_transactions":
                return {"data": [
                    _charge("bt1", 5000, 175, "evt-1", "a@x.com", "Workshop"),
                    _charge("bt2", 2000, 100, "evt-2", "b@x.com", "Mixer"),
                    {"id": "btref", "type": "refund", "amount": -500, "net": -500, "currency": "aud", "description": "refund"},
                    {"id": "btpo", "type": "payout", "amount": -6225, "net": -6225, "currency": "aud"},
                ], "has_more": False}
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=False)
        p = report["payouts"][0]
        self.assertEqual(p["deposit_cents"], 6225)  # 4825 + 1900 - 500
        brief = base64.b64decode(report["brief"]["content_base64"]).decode("utf-8")
        self.assertAlmostEqual(_sum_split_amounts(brief), _dollars(p["deposit_cents"]), places=2)

    def test_standalone_stripe_fee_folds_into_fee_not_refund(self):
        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [{"id": "po_F", "amount": 4795, "currency": "aud", "arrival_date": 1, "status": "paid"}], "has_more": False}
            if path == "/v1/balance_transactions":
                return {"data": [
                    _charge("bt1", 5000, 175, "evt-1", "a@x.com", "Workshop"),
                    {"id": "btfee", "type": "stripe_fee", "amount": -30, "net": -30, "currency": "aud", "description": "monthly fee"},
                    {"id": "btpo", "type": "payout", "amount": -4795, "net": -4795, "currency": "aud"},
                ], "has_more": False}
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=False)
        p = report["payouts"][0]
        self.assertEqual(p["stripe_fee_cents"], 205)  # 175 charge fee + 30 standalone
        self.assertEqual(p["refunds"], [])  # not listed as a refund
        self.assertEqual(p["deposit_cents"], 4795)
        brief = base64.b64decode(report["brief"]["content_base64"]).decode("utf-8")
        self.assertAlmostEqual(_sum_split_amounts(brief), _dollars(p["deposit_cents"]), places=2)

    def test_workbook_has_expected_sheets_and_values(self):
        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [{"id": "po_W", "amount": 4825, "currency": "aud", "arrival_date": 1_780_600_000, "status": "paid"}], "has_more": False}
            if path == "/v1/balance_transactions":
                return {"data": [_charge("bt1", 5000, 175, "evt-1", "a@x.com", "Workshop")], "has_more": False}
            raise AssertionError(path)

        from io import BytesIO
        from openpyxl import load_workbook

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=True)
        wb = load_workbook(BytesIO(base64.b64decode(report["workbook"]["content_base64"])))
        self.assertEqual(wb.sheetnames, ["Sales detail", "Payout summary"])
        summary = wb["Payout summary"]
        headers = [c.value for c in summary[1]]
        self.assertIn("Bank deposit", headers)
        deposit_col = headers.index("Bank deposit")
        self.assertEqual(summary[2][deposit_col].value, _dollars(report["payouts"][0]["deposit_cents"]))

    def test_injection_is_sanitized_in_brief_and_workbook(self):
        evil = '=HYPERLINK("http://evil")|pipe\nsecond'

        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [{"id": "po_X", "amount": 4825, "currency": "aud", "arrival_date": 1, "status": "paid"}], "has_more": False}
            if path == "/v1/balance_transactions":
                return {"data": [_charge("bt1", 5000, 175, "evt-1", "=cmd@x.com", evil)], "has_more": False}
            raise AssertionError(path)

        from io import BytesIO
        from openpyxl import load_workbook

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=True)

        brief = base64.b64decode(report["brief"]["content_base64"]).decode("utf-8")
        self.assertIn("\\|pipe", brief)  # pipe escaped in markdown
        # no brief table row contains a raw newline splitting the event name
        self.assertNotIn("|pipe\nsecond", brief)

        wb = load_workbook(BytesIO(base64.b64decode(report["workbook"]["content_base64"])))
        ws = wb["Sales detail"]
        headers = [c.value for c in ws[1]]
        event_cell = ws[2][headers.index("Event")].value
        buyer_cell = ws[2][headers.index("Buyer")].value
        self.assertTrue(event_cell.startswith("'="))  # formula neutralised
        self.assertTrue(buyer_cell.startswith("'="))

    def test_control_char_does_not_crash_workbook(self):
        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [{"id": "po_C", "amount": 4825, "currency": "aud", "arrival_date": 1, "status": "paid"}], "has_more": False}
            if path == "/v1/balance_transactions":
                return {"data": [_charge("bt1", 5000, 175, "evt-1", "a@x.com", "Bad\x07Bell\x0bVtab")], "has_more": False}
            raise AssertionError(path)

        from io import BytesIO
        from openpyxl import load_workbook

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=True)
        self.assertIn("workbook", report)  # produced, not degraded/crashed
        wb = load_workbook(BytesIO(base64.b64decode(report["workbook"]["content_base64"])))
        ws = wb["Sales detail"]
        headers = [c.value for c in ws[1]]
        self.assertEqual(ws[2][headers.index("Event")].value, "BadBellVtab")  # control chars stripped

    def test_multi_currency_is_not_cross_summed(self):
        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [
                    {"id": "po_aud", "amount": 4825, "currency": "aud", "arrival_date": 1, "status": "paid"},
                    {"id": "po_usd", "amount": 1900, "currency": "usd", "arrival_date": 2, "status": "paid"},
                ], "has_more": False}
            if path == "/v1/balance_transactions":
                if params["payout"] == "po_aud":
                    return {"data": [_charge("b1", 5000, 175, "evt-1", "a@x.com", "A")], "has_more": False}
                return {"data": [_charge("b2", 2000, 100, "evt-2", "b@x.com", "B")], "has_more": False}
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=False)
        self.assertEqual(report["currency_totals"]["AUD"]["deposit"], 48.25)
        self.assertEqual(report["currency_totals"]["USD"]["deposit"], 19.00)

    def test_unmatched_charges_get_separate_buckets(self):
        def handler(path, params):
            if path == "/v1/payouts":
                return {"data": [{"id": "po_U", "amount": 4800, "currency": "aud", "arrival_date": 1, "status": "paid"}], "has_more": False}
            if path == "/v1/balance_transactions":
                return {"data": [
                    _charge("bt1", 2500, 100, "", "a@x.com", "Mystery A"),
                    _charge("bt2", 2500, 100, "", "b@x.com", "Mystery B"),
                ], "has_more": False}
            raise AssertionError(path)

        since, until = self._window()
        report = self._service(handler).build_report(since=since, until=until, include_workbook=False)
        p = report["payouts"][0]
        self.assertEqual(len(p["events"]), 2)  # not collapsed into one __unknown__ bucket
        self.assertEqual(report["unmatched_charge_count"], 2)


class ReconciliationReportViewTests(APITestCase):
    def setUp(self):
        self.url = reverse("reconciliation_report")
        self.admin = "URADMIN"
        self.committee = "URCOMMITTEE"
        self.portfolio = "URPORTFOLIO"
        self.partner = "URPARTNER"
        self.inactive = "URINACTIVE"
        PointsAdmin.objects.create(slack_user_id=self.admin, role="admin", is_active=True)
        PointsAdmin.objects.create(slack_user_id=self.committee, role="committee", is_active=True)
        PointsAdmin.objects.create(slack_user_id=self.portfolio, role="portfolio_lead", is_active=True)
        PointsAdmin.objects.create(slack_user_id=self.partner, role="partner", is_active=True)
        PointsAdmin.objects.create(slack_user_id=self.inactive, role="admin", is_active=False)

    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_full_admins_allowed_partner_and_others_denied(self, _perm):
        class FakeService:
            def build_report(self, **kwargs):
                return {"payouts": [], "payout_count": 0}

        with patch("integrations.api_views_reconciliation.ReconciliationReportService", return_value=FakeService()):
            for slack_id in [self.admin, self.committee, self.portfolio]:
                resp = self.client.get(self.url, {"slack_user_id": slack_id})
                self.assertEqual(resp.status_code, status.HTTP_200_OK, slack_id)

        denied = SimpleNamespace(build_report=lambda **kw: self.fail("should not call Stripe"))
        with patch("integrations.api_views_reconciliation.ReconciliationReportService", return_value=denied):
            for slack_id in ["UNOBODY", self.partner, self.inactive]:
                resp = self.client.get(self.url, {"slack_user_id": slack_id})
                self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN, slack_id)

    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_requires_slack_user_id(self, _perm):
        resp = self.client.get(self.url, {})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    @override_settings(STRIPE_SECRET_KEY="")
    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_missing_stripe_key_returns_configuration_error(self, _perm):
        resp = self.client.get(self.url, {"slack_user_id": self.admin})
        self.assertEqual(resp.status_code, status.HTTP_503_SERVICE_UNAVAILABLE)
        self.assertIn("STRIPE_SECRET_KEY", resp.data["error"])

    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_window_params_default_and_explicit(self, _perm):
        captured = []

        class FakeService:
            def build_report(self, **kwargs):
                captured.append(kwargs)
                return {"payouts": []}

        with patch("integrations.api_views_reconciliation.ReconciliationReportService", return_value=FakeService()):
            resp = self.client.get(self.url, {"slack_user_id": self.admin})
            self.assertEqual(resp.status_code, status.HTTP_200_OK)
            # default 30-day window
            span = captured[0]["until"] - captured[0]["since"]
            self.assertEqual(span.days, 30)

            captured.clear()
            resp = self.client.get(
                self.url,
                {"slack_user_id": self.admin, "since": "2026-06-01", "until": "2026-06-30", "include_workbook": "false"},
            )
            self.assertEqual(resp.status_code, status.HTTP_200_OK)
            self.assertEqual(captured[0]["since"].date().isoformat(), "2026-06-01")
            self.assertEqual(captured[0]["include_workbook"], False)

    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_window_validation_errors(self, _perm):
        # since after until -> 400 (no service call needed)
        resp = self.client.get(self.url, {"slack_user_id": self.admin, "since": "2026-06-30", "until": "2026-06-01"})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        # bad date format -> 400
        resp = self.client.get(self.url, {"slack_user_id": self.admin, "since": "not-a-date"})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_days_window_is_capped(self, _perm):
        captured = []

        class FakeService:
            def build_report(self, **kwargs):
                captured.append(kwargs)
                return {"payouts": []}

        with patch("integrations.api_views_reconciliation.ReconciliationReportService", return_value=FakeService()):
            resp = self.client.get(self.url, {"slack_user_id": self.admin, "days": "500"})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual((captured[0]["until"] - captured[0]["since"]).days, 92)  # MAX_WINDOW_DAYS

    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_stripe_errors_map_to_status(self, _perm):
        class RateLimited:
            def build_report(self, **kwargs):
                raise StripeAPIError("rate limited", status_code=429)

        with patch("integrations.api_views_reconciliation.ReconciliationReportService", return_value=RateLimited()):
            resp = self.client.get(self.url, {"slack_user_id": self.admin})
            self.assertEqual(resp.status_code, status.HTTP_429_TOO_MANY_REQUESTS)

        class Bad:
            def build_report(self, **kwargs):
                raise StripeAPIError("boom", status_code=500)

        with patch("integrations.api_views_reconciliation.ReconciliationReportService", return_value=Bad()):
            resp = self.client.get(self.url, {"slack_user_id": self.admin})
            self.assertEqual(resp.status_code, status.HTTP_502_BAD_GATEWAY)

    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_malformed_days_is_rejected_even_with_since(self, _perm):
        # days is validated unconditionally, so a bad value 400s even when since wins
        resp = self.client.get(
            self.url,
            {"slack_user_id": self.admin, "since": "2026-06-01", "until": "2026-06-10", "days": "abc"},
        )
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    @patch("core.permissions.HasRooApiKey.has_permission", return_value=True)
    def test_exactly_max_window_is_allowed(self, _perm):
        # 2026-04-01 .. 2026-07-02 is exactly 92 calendar days; must not 400
        captured = []

        class FakeService:
            def build_report(self, **kwargs):
                captured.append(kwargs)
                return {"payouts": []}

        with patch("integrations.api_views_reconciliation.ReconciliationReportService", return_value=FakeService()):
            resp = self.client.get(self.url, {"slack_user_id": self.admin, "since": "2026-04-01", "until": "2026-07-02"})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual((captured[0]["until"].date() - captured[0]["since"].date()).days, 92)
