"""
Luma -> Stripe -> Xero reconciliation service.

Builds a monthly reconciliation report for MLAI ticket sales. The pull is
**payout-driven** and **Stripe is the spine**: for every Stripe payout that
settled in the window we pull the balance transactions inside it, so each
payout's charge nets tie to the exact bank deposit.

Every Luma ticket charge carries its own attribution in Stripe metadata
(`event_api_id`, buyer `email`) and the event name in the charge `description`
(validated against live data 2026-07-01 — see the roo repo's
Luma-Stripe-Reconcile/PHASE-0.2-FINDINGS.md). So the report is assembled from
Stripe alone; Luma is only an optional enrichment and is not required here.

Output (one JSON payload):
- ``payouts`` — per payout: net (= bank deposit), gross/fee split, per-event
  breakdown, per-charge rows, refunds, and any tie-out warnings.
- ``brief`` — a base64 markdown brief for Claude Cowork to reconcile in Xero.
- ``workbook`` — an optional base64 .xlsx audit workbook (Sales detail + Payout
  summary). Omitted gracefully if openpyxl is unavailable.

Read-only. Never writes to Stripe. Buyer emails are PII — the calling view gates
this on Points Admin.
"""
from __future__ import annotations

import base64
import io
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import requests
from django.conf import settings


STRIPE_API_BASE_URL = "https://api.stripe.com"
DEFAULT_STRIPE_API_VERSION = "2026-02-25.clover"
_STRIPE_API_VERSION_RE = re.compile(r"^\d{4}-\d{2}-\d{2}(?:\.[A-Za-z0-9_-]+)?$")

# Stripe balance-transaction types that represent processing fees (folded into
# the Stripe Fees line rather than listed as refunds/adjustments).
STRIPE_FEE_TYPES = {"stripe_fee", "application_fee", "tax_fee"}
REFUND_TYPES = {"refund", "payment_refund"}
REVENUE_TYPES = {"charge", "payment"}

# openpyxl rejects most C0 control chars (it allows tab \x09, LF \x0a, CR \x0d).
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")

# Xero mapping labels used in the Cowork brief. The concrete Xero account codes
# and tracking-category options are wired up in Phase 2 (see the build plan);
# here they are human-readable placeholders in the brief.
TICKET_INCOME_ACCOUNT = "Ticket Sales"
STRIPE_FEE_ACCOUNT = "Stripe Fees"
TAX_RATE_LABEL = "account default"


class StripeConfigurationError(RuntimeError):
    """Raised when the backend is missing Stripe configuration."""


class StripeAPIError(RuntimeError):
    """Raised when Stripe rejects or fails an API request."""

    def __init__(self, message: str, *, status_code: Optional[int] = None):
        super().__init__(message)
        self.status_code = status_code


class ReconciliationMappingSource:
    LUMA_EVENT = "luma_event"
    STRIPE_INVOICE = "stripe_invoice"
    STRIPE_METADATA = "stripe_metadata"
    UNATTRIBUTED = "unattributed"


def _dollars(cents: int) -> float:
    return round((cents or 0) / 100.0, 2)


def _clean(value: Any) -> str:
    """Drop xlsx-illegal control chars from a string value."""
    return _ILLEGAL_XLSX_CHARS.sub("", "" if value is None else str(value))


def _safe_cell(value: Any) -> Any:
    """Neutralise spreadsheet formula injection for string cells.

    Buyer-controlled text (event name from Stripe description, buyer email from
    metadata) is written into the audit xlsx. A value starting with = + - @
    (or tab/CR/LF) would be interpreted as a live formula when the workbook is
    opened, so prefix it with a single quote. Non-strings pass through.
    """
    if not isinstance(value, str):
        return value
    s = _clean(value)
    if s[:1] in _FORMULA_PREFIXES:
        return "'" + s
    return s


def _md_cell(value: Any) -> str:
    """Escape a value for safe interpolation into a markdown table cell."""
    s = _clean(value).replace("\r", " ").replace("\n", " ").replace("|", "\\|")
    return s


def resolve_stripe_api_version(value: Any) -> str:
    """Return a valid Stripe-Version header, safely repairing malformed env values.

    A production env file once contained a literal Python settings expression.
    Sending that text as a header makes every Stripe request fail with a 400, so
    malformed values use the pinned application default instead.
    """
    candidate = str(value or "").strip()
    return candidate if _STRIPE_API_VERSION_RE.fullmatch(candidate) else DEFAULT_STRIPE_API_VERSION


class ReconciliationReportService:
    """Pull Stripe payouts + charges and build the reconciliation report."""

    def __init__(
        self,
        *,
        stripe_api_key: Optional[str] = None,
        stripe_api_version: Optional[str] = None,
        base_url: Optional[str] = None,
        session: Optional[requests.Session] = None,
        timeout: Any = (3, 20),
    ):
        raw_key = (
            stripe_api_key
            if stripe_api_key is not None
            else getattr(settings, "STRIPE_SECRET_KEY", None)
        )
        self.stripe_api_key = str(raw_key or "").strip()
        configured_version = (
            stripe_api_version
            if stripe_api_version is not None
            else getattr(settings, "STRIPE_API_VERSION", DEFAULT_STRIPE_API_VERSION)
        )
        self.stripe_api_version = resolve_stripe_api_version(configured_version)
        self.base_url = str(base_url or STRIPE_API_BASE_URL).rstrip("/")
        self.session = session or requests.Session()
        self.timeout = timeout

    # ---- public API ------------------------------------------------------

    def build_report(
        self,
        *,
        since: datetime,
        until: datetime,
        include_workbook: bool = True,
    ) -> Dict[str, Any]:
        if not self.stripe_api_key:
            raise StripeConfigurationError("STRIPE_SECRET_KEY is not configured on mlai-backend.")

        start_unix = int(since.timestamp())
        end_unix = int(until.timestamp())

        payouts = self._list(
            "/v1/payouts",
            {"arrival_date[gte]": start_unix, "arrival_date[lt]": end_unix, "status": "paid"},
        )

        payout_reports: List[Dict[str, Any]] = []
        sales_rows: List[Dict[str, Any]] = []
        unmatched: List[Dict[str, Any]] = []
        currency_totals: Dict[str, Dict[str, Any]] = {}

        for payout in sorted(payouts, key=lambda p: p.get("arrival_date") or 0):
            report, rows = self._build_payout_report(payout)
            payout_reports.append(report)
            sales_rows.extend(rows)
            for row in rows:
                if row.get("source_type") == ReconciliationMappingSource.UNATTRIBUTED:
                    unmatched.append(row)

            ccy = report["currency"]
            bucket = currency_totals.setdefault(
                ccy,
                {"payouts": 0, "gross_cents": 0, "stripe_fee_cents": 0, "deposit_cents": 0},
            )
            bucket["payouts"] += 1
            bucket["gross_cents"] += report["gross_cents"]
            bucket["stripe_fee_cents"] += report["stripe_fee_cents"]
            bucket["deposit_cents"] += report["deposit_cents"]

        result: Dict[str, Any] = {
            "window": {
                "since": since.astimezone(timezone.utc).isoformat(),
                "until": until.astimezone(timezone.utc).isoformat(),
            },
            "currency_totals": {
                ccy: {
                    "payouts": b["payouts"],
                    "gross": _dollars(b["gross_cents"]),
                    "stripe_fee": _dollars(b["stripe_fee_cents"]),
                    "deposit": _dollars(b["deposit_cents"]),
                }
                for ccy, b in sorted(currency_totals.items())
            },
            "payout_count": len(payout_reports),
            "charge_count": len(sales_rows),
            "unmatched_charge_count": len(unmatched),
            "payouts": payout_reports,
        }

        brief_md = self._render_brief(payout_reports, result)
        result["brief"] = {
            "filename": self._brief_filename(since, until),
            "content_base64": base64.b64encode(brief_md.encode("utf-8")).decode("ascii"),
            "content_type": "text/markdown",
        }

        if include_workbook:
            workbook = self._render_workbook(sales_rows, payout_reports)
            if workbook is None:
                result["workbook_error"] = "openpyxl is not installed; xlsx workbook was skipped."
            else:
                result["workbook"] = {
                    "filename": self._brief_filename(since, until).replace(".md", ".xlsx"),
                    "content_base64": base64.b64encode(workbook).decode("ascii"),
                    "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }

        return result

    def build_payout(self, payout_id: str) -> Dict[str, Any]:
        """Build one payout ledger for a webhook-triggered incremental sync."""
        if not self.stripe_api_key:
            raise StripeConfigurationError("STRIPE_SECRET_KEY is not configured on mlai-backend.")
        normalized_id = str(payout_id or "").strip()
        if not normalized_id.startswith("po_"):
            raise StripeAPIError("Invalid Stripe payout id.")
        payout = self._get(f"/v1/payouts/{normalized_id}", {})
        if str(payout.get("status") or "") != "paid":
            raise StripeAPIError("Stripe payout is not paid yet.")
        report, rows = self._build_payout_report(payout)
        return {
            "payout_count": 1,
            "charge_count": len(rows),
            "unmatched_charge_count": sum(
                1 for row in rows if row.get("source_type") == ReconciliationMappingSource.UNATTRIBUTED
            ),
            "payouts": [report],
        }

    # ---- payout assembly -------------------------------------------------

    def _build_payout_report(self, payout: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        payout_id = str(payout.get("id") or "")
        currency = str(payout.get("currency") or "").upper()
        arrival = _unix_to_date(payout.get("arrival_date"))
        txns = self._list(
            "/v1/balance_transactions",
            {"payout": payout_id, "expand[]": "data.source"},
        )

        sales_rows: List[Dict[str, Any]] = []
        refunds: List[Dict[str, Any]] = []
        groups: Dict[str, Dict[str, Any]] = {}
        gross_cents = fee_cents = standalone_fee_cents = charge_net_cents = non_payout_net = 0

        for txn in txns:
            ttype = str(txn.get("type") or "")
            if ttype == "payout":
                continue
            non_payout_net += int(txn.get("net") or 0)

            if ttype in REVENUE_TYPES:
                source = txn.get("source") if isinstance(txn.get("source"), dict) else {}
                attribution = self._revenue_attribution(txn, source)
                metadata = attribution.pop("metadata")
                row_gross = int(txn.get("amount") or 0)
                row_fee = int(txn.get("fee") or 0)
                row_net = int(txn.get("net") or 0)
                gross_cents += row_gross
                fee_cents += row_fee
                charge_net_cents += row_net
                row = {
                    "charge_id": str(source.get("id") or txn.get("id") or ""),
                    "event_api_id": attribution["event_api_id"],
                    "event_name": attribution["source_label"],
                    **attribution,
                    "buyer_email": str(metadata.get("email") or (source.get("billing_details") or {}).get("email") or "").strip(),
                    "gross_cents": row_gross,
                    "stripe_fee_cents": row_fee,
                    "net_cents": row_net,
                    "currency": str(txn.get("currency") or currency).upper(),
                    "created": _unix_to_iso(source.get("created") or txn.get("created")),
                    "payout_id": payout_id,
                    "payout_arrival": arrival,
                }
                sales_rows.append(row)
                key = f"{row['source_type']}:{row['source_id']}"
                bucket = groups.setdefault(
                    key,
                    {
                        "event_api_id": row["event_api_id"],
                        "event_name": row["source_label"],
                        "source_type": row["source_type"],
                        "source_id": row["source_id"],
                        "source_label": row["source_label"],
                        "project_name": row.get("project_name", ""),
                        "ticket_count": 0,
                        "gross_cents": 0,
                        "stripe_fee_cents": 0,
                        "stripe_invoice_ids": set(),
                        "stripe_invoice_payment_ids": set(),
                        "stripe_payment_intent_ids": set(),
                        "stripe_product_ids": set(),
                        "stripe_subscription_ids": set(),
                    },
                )
                bucket["ticket_count"] += 1
                bucket["gross_cents"] += row_gross
                bucket["stripe_fee_cents"] += row_fee
                for field in (
                    "stripe_invoice_ids",
                    "stripe_invoice_payment_ids",
                    "stripe_payment_intent_ids",
                    "stripe_product_ids",
                    "stripe_subscription_ids",
                ):
                    values = row.get(field) or []
                    bucket[field].update(str(value) for value in values if str(value))
            elif ttype in STRIPE_FEE_TYPES or txn.get("reporting_category") == "fee":
                standalone_fee = -int(txn.get("net") or 0)
                fee_cents += standalone_fee
                standalone_fee_cents += standalone_fee
            elif ttype in REFUND_TYPES:
                source = txn.get("source") if isinstance(txn.get("source"), dict) else {}
                refunds.append(
                    {
                        "id": str(txn.get("id") or ""),
                        "type": ttype,
                        "amount_cents": int(txn.get("amount") or 0),
                        "net_cents": int(txn.get("net") or 0),
                        "currency": str(txn.get("currency") or currency).upper(),
                        "description": str(txn.get("description") or source.get("description") or ""),
                        **self._refund_attribution(source),
                    }
                )
            else:
                refunds.append(
                    {
                        "id": str(txn.get("id") or ""),
                        "type": ttype,
                        "amount_cents": int(txn.get("amount") or 0),
                        "net_cents": int(txn.get("net") or 0),
                        "currency": str(txn.get("currency") or currency).upper(),
                        "description": str(txn.get("description") or ""),
                        "source_type": ReconciliationMappingSource.UNATTRIBUTED,
                        "source_id": str(txn.get("id") or ""),
                        "source_label": str(txn.get("description") or ttype),
                        "event_api_id": "",
                    }
                )

        refund_net_cents = sum(int(r.get("net_cents") or 0) for r in refunds)
        payout_amount = int(payout.get("amount") or 0)
        warnings: List[str] = []
        if non_payout_net != payout_amount:
            warnings.append(
                "Tie-out mismatch: sum of transaction nets "
                f"({_dollars(non_payout_net)}) != payout amount ({_dollars(payout_amount)})."
            )
        if any(r["source_type"] == ReconciliationMappingSource.UNATTRIBUTED for r in sales_rows):
            warnings.append(
                "Some payments have no Luma event_api_id and could not be attributed — "
                "a source mapping is required."
            )
        if refunds:
            warnings.append(f"{len(refunds)} non-charge transaction(s) (refunds/adjustments) in this payout.")
        ordered_groups = sorted(groups.values(), key=lambda item: item["event_name"])
        for group in ordered_groups:
            for field in (
                "stripe_invoice_ids",
                "stripe_invoice_payment_ids",
                "stripe_payment_intent_ids",
                "stripe_product_ids",
                "stripe_subscription_ids",
            ):
                group[field] = sorted(group[field])
        report = {
            "payout_id": payout_id,
            "arrival_date": arrival,
            "currency": currency,
            "payout_amount_cents": payout_amount,
            "deposit_cents": payout_amount,
            "gross_cents": gross_cents,
            "stripe_fee_cents": fee_cents,
            "standalone_fee_cents": standalone_fee_cents,
            "charge_net_cents": charge_net_cents,
            "refund_net_cents": refund_net_cents,
            "charge_count": len(sales_rows),
            "events": ordered_groups,
            "revenue_groups": ordered_groups,
            "charges": sales_rows,
            "refunds": refunds,
            "warnings": warnings,
        }
        return report, sales_rows

    def _revenue_attribution(self, txn: Dict[str, Any], source: Dict[str, Any]) -> Dict[str, Any]:
        metadata = source.get("metadata") if isinstance(source.get("metadata"), dict) else {}
        event_api_id = str(metadata.get("event_api_id") or "").strip()
        source_object_id = str(source.get("id") or txn.get("id") or "").strip()
        description = str(source.get("description") or txn.get("description") or "").strip()
        payment_intent_id = self._stripe_id(source.get("payment_intent"))
        if not payment_intent_id and source_object_id.startswith("pi_"):
            payment_intent_id = source_object_id
        invoice_id = self._stripe_id(source.get("invoice"))
        product_ids = self._stripe_product_ids(source, metadata)
        subscription_ids = self._stripe_subscription_ids(source, metadata)
        lineage = {
            "stripe_balance_transaction_id": str(txn.get("id") or ""),
            "stripe_source_object_id": source_object_id,
            "stripe_invoice_ids": [invoice_id] if invoice_id else [],
            "stripe_invoice_payment_ids": [],
            "stripe_payment_intent_ids": [payment_intent_id] if payment_intent_id else [],
            "stripe_product_ids": sorted(product_ids),
            "stripe_subscription_ids": sorted(subscription_ids),
        }
        if event_api_id:
            return {"source_type": ReconciliationMappingSource.LUMA_EVENT, "source_id": event_api_id, "source_label": description or event_api_id, "event_api_id": event_api_id, "project_name": "", "metadata": metadata, **lineage}
        if metadata.get("points_purchase_id") or metadata.get("pack_id") or metadata.get("points"):
            return {"source_type": ReconciliationMappingSource.STRIPE_METADATA, "source_id": "roo_points", "source_label": description or "Roo Points", "event_api_id": "", "project_name": "Roo Points", "metadata": metadata, **lineage}
        invoice, invoice_payment_id = self._invoice_for_payment(source)
        if invoice:
            invoice_id = str(invoice.get("id") or source_object_id)
            lines = invoice.get("lines") if isinstance(invoice.get("lines"), dict) else {}
            descriptions = [str(line.get("description") or "").strip() for line in lines.get("data", []) if isinstance(line, dict) and str(line.get("description") or "").strip()]
            label = descriptions[0] if descriptions else description or invoice_id
            lineage["stripe_invoice_ids"] = [invoice_id]
            lineage["stripe_invoice_payment_ids"] = (
                [invoice_payment_id] if invoice_payment_id else []
            )
            lineage["stripe_product_ids"] = sorted(
                set(lineage["stripe_product_ids"]) | self._invoice_product_ids(invoice)
            )
            lineage["stripe_subscription_ids"] = sorted(
                set(lineage["stripe_subscription_ids"])
                | self._invoice_subscription_ids(invoice)
            )
            return {"source_type": ReconciliationMappingSource.STRIPE_INVOICE, "source_id": invoice_id, "source_label": label, "event_api_id": "", "project_name": label, "stripe_invoice_id": invoice_id, "metadata": metadata, **lineage}
        if lineage["stripe_product_ids"]:
            product_id = lineage["stripe_product_ids"][0]
            return {"source_type": "stripe_product", "source_id": product_id, "source_label": description or product_id, "event_api_id": "", "project_name": "", "metadata": metadata, **lineage}
        return {"source_type": ReconciliationMappingSource.UNATTRIBUTED, "source_id": source_object_id, "source_label": description or "(unattributed Stripe payment)", "event_api_id": "", "project_name": "", "metadata": metadata, **lineage}

    @staticmethod
    def _stripe_id(value: Any) -> str:
        return str(value.get("id") if isinstance(value, dict) else value or "").strip()

    @classmethod
    def _stripe_product_ids(cls, source: Dict[str, Any], metadata: Dict[str, Any]) -> set[str]:
        price = source.get("price") if isinstance(source.get("price"), dict) else {}
        values = {
            cls._stripe_id(source.get("product")),
            cls._stripe_id(price.get("product")),
            str(metadata.get("product_id") or "").strip(),
        }
        return {value for value in values if value.startswith("prod_")}

    @staticmethod
    def _stripe_subscription_ids(source: Dict[str, Any], metadata: Dict[str, Any]) -> set[str]:
        values = {
            ReconciliationReportService._stripe_id(source.get("subscription")),
            str(metadata.get("subscription_id") or "").strip(),
        }
        return {value for value in values if value.startswith("sub_")}

    @classmethod
    def _invoice_product_ids(cls, invoice: Dict[str, Any]) -> set[str]:
        lines = invoice.get("lines") if isinstance(invoice.get("lines"), dict) else {}
        values: set[str] = set()
        for line in lines.get("data") or []:
            if not isinstance(line, dict):
                continue
            price = line.get("price") if isinstance(line.get("price"), dict) else {}
            pricing = line.get("pricing") if isinstance(line.get("pricing"), dict) else {}
            price_details = (
                pricing.get("price_details")
                if isinstance(pricing.get("price_details"), dict)
                else {}
            )
            values.update({
                cls._stripe_id(line.get("product")),
                cls._stripe_id(price.get("product")),
                cls._stripe_id(price_details.get("product")),
            })
        return {value for value in values if value.startswith("prod_")}

    @classmethod
    def _invoice_subscription_ids(cls, invoice: Dict[str, Any]) -> set[str]:
        parent = invoice.get("parent") if isinstance(invoice.get("parent"), dict) else {}
        details = (
            parent.get("subscription_details")
            if isinstance(parent.get("subscription_details"), dict)
            else {}
        )
        values = {
            cls._stripe_id(invoice.get("subscription")),
            cls._stripe_id(details.get("subscription")),
        }
        return {value for value in values if value.startswith("sub_")}

    def _invoice_for_payment(self, source: Dict[str, Any]) -> tuple[Dict[str, Any], str]:
        raw_invoice = source.get("invoice")
        if isinstance(raw_invoice, dict):
            return raw_invoice, ""
        invoice_id = str(raw_invoice or "").strip()
        raw_intent = source.get("payment_intent")
        intent_id = str(raw_intent.get("id") if isinstance(raw_intent, dict) else raw_intent or "").strip()
        if not intent_id and str(source.get("id") or "").startswith("pi_"):
            intent_id = str(source["id"])
        invoice_payment_id = ""
        if not invoice_id and intent_id:
            payments = self._list("/v1/invoice_payments", {"payment[type]": "payment_intent", "payment[payment_intent]": intent_id})
            if payments:
                invoice_payment_id = str(payments[0].get("id") or "").strip()
                raw_invoice = payments[0].get("invoice")
                invoice_id = str(raw_invoice.get("id") if isinstance(raw_invoice, dict) else raw_invoice or "").strip()
        return (
            self._get(f"/v1/invoices/{invoice_id}", {}) if invoice_id else {},
            invoice_payment_id,
        )

    def _refund_attribution(self, source: Dict[str, Any]) -> Dict[str, Any]:
        raw_intent = source.get("payment_intent")
        if isinstance(raw_intent, dict):
            intent = raw_intent
        else:
            intent_id = str(raw_intent or "").strip()
            intent = self._get(f"/v1/payment_intents/{intent_id}", {}) if intent_id else {}
        metadata = intent.get("metadata") if isinstance(intent.get("metadata"), dict) else {}
        event_id = str(metadata.get("event_api_id") or "").strip()
        intent_id = str(intent.get("id") or "").strip()
        lineage = {
            "stripe_refund_id": str(source.get("id") or ""),
            "stripe_payment_intent_id": intent_id,
            "stripe_invoice_id": self._stripe_id(intent.get("invoice")),
            "stripe_charge_id": self._stripe_id(source.get("charge")),
        }
        if event_id:
            return {"source_type": ReconciliationMappingSource.LUMA_EVENT, "source_id": event_id, "source_label": str(intent.get("description") or event_id), "event_api_id": event_id, **lineage}
        return {"source_type": ReconciliationMappingSource.UNATTRIBUTED, "source_id": str(intent.get("id") or source.get("id") or ""), "source_label": str(intent.get("description") or source.get("description") or ""), "event_api_id": "", **lineage}

    # ---- Stripe HTTP -----------------------------------------------------

    def _list(self, path: str, params: Dict[str, Any]) -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        starting_after: Optional[str] = None
        while True:
            page_params = dict(params, limit=100)
            if starting_after:
                page_params["starting_after"] = starting_after
            page = self._get(path, page_params)
            rows = page.get("data", []) if isinstance(page, dict) else []
            out.extend(rows)
            if not page.get("has_more") or not rows:
                break
            last_id = rows[-1].get("id")
            if not last_id:
                break
            starting_after = last_id
        return out

    def _get(self, path: str, params: Dict[str, Any]) -> Dict[str, Any]:
        url = f"{self.base_url}{path}"
        headers = {
            "Authorization": f"Bearer {self.stripe_api_key}",
            "Stripe-Version": self.stripe_api_version,
        }
        try:
            response = self.session.get(url, headers=headers, params=params, timeout=self.timeout)
        except requests.RequestException as exc:
            raise StripeAPIError("Unable to reach Stripe.") from exc

        status_code = getattr(response, "status_code", None)
        if status_code in (401, 403):
            raise StripeAPIError("Stripe rejected the configured API key.", status_code=status_code)
        if status_code == 429:
            raise StripeAPIError("Stripe rate-limited the reconciliation request.", status_code=status_code)

        try:
            response.raise_for_status()
        except requests.RequestException as exc:
            raise StripeAPIError("Stripe returned an error.", status_code=status_code) from exc

        try:
            payload = response.json()
        except ValueError as exc:
            raise StripeAPIError("Stripe returned an invalid JSON response.", status_code=status_code) from exc
        return payload if isinstance(payload, dict) else {}

    # ---- rendering -------------------------------------------------------

    @staticmethod
    def _brief_filename(since: datetime, until: datetime) -> str:
        a = since.astimezone(timezone.utc).date().isoformat()
        b = until.astimezone(timezone.utc).date().isoformat()
        return f"reconciliation-{a}-to-{b}.md"

    def _render_brief(self, payouts: List[Dict[str, Any]], summary: Dict[str, Any]) -> str:
        L: List[str] = []
        L.append("# Stripe payout reconciliation brief")
        L.append("")
        L.append("**For review.** Reconcile each Stripe deposit below in Xero")
        L.append("bank feed. **Only these Stripe/Luma deposits — leave every other bank")
        L.append("line (transfers, PayID, etc.) untouched.**")
        L.append("")
        L.append("For each: open the bank line -> **Create** -> fill Who / Why, then")
        L.append("**Add details** to split by source. Each split line sets **What** (account),")
        L.append("**Event Name**, **Project Name**, **Amount**, **Tax Rate**. The split total")
        L.append("must equal the bank line to the cent. **A human clicks OK to confirm.**")
        L.append("")
        L.append("> Tax Rate is left as the income account's default — confirm it in Xero.")
        L.append("")

        L.append("## Deposits to reconcile")
        L.append("")
        L.append("| Payout | Arrived | Bank deposit | Sources | Payments |")
        L.append("|---|---|---|---|---|")
        for p in payouts:
            evs = ", ".join(_md_cell(e["event_name"]) for e in p["events"]) or "—"
            L.append(
                f"| `{p['payout_id']}` | {p['arrival_date']} | "
                f"{p['currency']} {_dollars(p['deposit_cents']):,.2f} | {evs} | {p['charge_count']} |"
            )
        if summary.get("unmatched_charge_count"):
            L.append("")
            L.append(
                f"> ⚠ {summary['unmatched_charge_count']} payment(s) could not be attributed "
                "from Stripe/Luma metadata — map them manually (marked below)."
            )
        L.append("")
        L.append("---")
        L.append("")

        for p in payouts:
            ccy = p["currency"]
            L.append(f"## {p['payout_id']} — {ccy} {_dollars(p['deposit_cents']):,.2f} received {p['arrival_date']}")
            L.append("")
            L.append(
                f"**Match the bank line:** Received **{ccy} {_dollars(p['deposit_cents']):,.2f}** "
                f"on/around **{p['arrival_date']}** (payer: Stripe)."
            )
            L.append("")
            L.append("- **Who:** Stripe Payments")
            events_list = ", ".join(_md_cell(e["event_name"]) for e in p["events"]) or "—"
            L.append(f"- **Why:** Stripe sales — {events_list} — {p['charge_count']} payment(s) — payout {p['payout_id']}")
            L.append("")
            L.append("**Create → Add details (split lines):**")
            L.append("")
            L.append("| What (account) | Event Name | Project Name | Amount | Tax Rate |")
            L.append("|---|---|---|---|---|")
            for e in p["events"]:
                name = _md_cell(e["event_name"])
                unmatched = e.get("source_type") == ReconciliationMappingSource.UNATTRIBUTED
                event_label = name if e.get("source_type") == ReconciliationMappingSource.LUMA_EVENT else "—"
                project_label = _md_cell(e.get("project_name") or "—")
                if unmatched:
                    project_label = f"⚠ map — {name}"
                L.append(
                    f"| Approved revenue/clearing account | {event_label} | {project_label} | "
                    f"{_dollars(e['gross_cents']):,.2f} | {TAX_RATE_LABEL} |"
                )
            if p["stripe_fee_cents"]:
                L.append(
                    f"| {STRIPE_FEE_ACCOUNT} | — | — | "
                    f"-{_dollars(p['stripe_fee_cents']):,.2f} | {TAX_RATE_LABEL} |"
                )
            for r in p["refunds"]:
                desc = _md_cell(r.get("description") or r.get("type") or "refund/adjustment")
                L.append(
                    f"| Refund/adjustment — {desc} | — | — | "
                    f"{_dollars(r.get('net_cents', 0)):,.2f} | {TAX_RATE_LABEL} |"
                )
            L.append(
                f"| **TOTAL — must equal the bank line** | | | "
                f"**{_dollars(p['deposit_cents']):,.2f}** | |"
            )
            L.append("")
            if p["warnings"]:
                for w in p["warnings"]:
                    L.append(f"> ⚠ {w}")
                L.append("")
            L.append("<details><summary>Buyers in this payout (audit — not entered per line)</summary>")
            L.append("")
            for c in p["charges"]:
                L.append(
                    f"- {_md_cell(c['buyer_email'] or '(no email)')} — {_md_cell(c['event_name'])} — "
                    f"{ccy} {_dollars(c['gross_cents']):,.2f} — {c['created']}"
                )
            L.append("")
            L.append("</details>")
            L.append("")
            L.append("> Cowork fills the form; **a human clicks OK to confirm the reconcile.**")
            L.append("")

        return "\n".join(L)

    def _render_workbook(
        self, sales_rows: List[Dict[str, Any]], payouts: List[Dict[str, Any]]
    ) -> Optional[bytes]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return None

        try:
            return self._build_workbook_bytes(
                Workbook, Alignment, Font, PatternFill, get_column_letter, sales_rows, payouts
            )
        except Exception:
            # Workbook is documented as optional — degrade to the workbook_error
            # path rather than 500-ing the whole report on a rendering fault.
            return None

    def _build_workbook_bytes(
        self, Workbook, Alignment, Font, PatternFill, get_column_letter, sales_rows, payouts
    ) -> bytes:
        head_fill = PatternFill("solid", fgColor="1F3864")
        head_font = Font(bold=True, color="FFFFFF")

        wb = Workbook()

        def render(ws, columns, data):
            ws.append([label for _, label, _ in columns])
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(1, col_idx)
                cell.fill, cell.font = head_fill, head_font
                cell.alignment = Alignment(horizontal="center")
            for record in data:
                ws.append([_safe_cell(fn(record)) for _, _, fn in columns])
            for col_idx, (_, label, fn) in enumerate(columns, start=1):
                letter = get_column_letter(col_idx)
                values = [len(str(label))] + [len(str(fn(r))) for r in data]
                ws.column_dimensions[letter].width = min(max(max(values) + 3, 11), 42)
            ws.freeze_panes = "A2"

        sales_cols = [
            ("created", "Date bought (UTC)", lambda r: r["created"]),
            ("event_name", "Event", lambda r: r["event_name"]),
            ("event_api_id", "Event id", lambda r: r["event_api_id"]),
            ("buyer_email", "Buyer", lambda r: r["buyer_email"]),
            ("gross", "Gross", lambda r: _dollars(r["gross_cents"])),
            ("stripe_fee", "Stripe fee", lambda r: _dollars(r["stripe_fee_cents"])),
            ("net", "Net", lambda r: _dollars(r["net_cents"])),
            ("currency", "Ccy", lambda r: r["currency"]),
            ("charge_id", "Stripe charge", lambda r: r["charge_id"]),
            ("payout_id", "Payout", lambda r: r["payout_id"]),
            ("payout_arrival", "Payout date", lambda r: r["payout_arrival"]),
        ]
        ws1 = wb.active
        ws1.title = "Sales detail"
        render(ws1, sales_cols, sales_rows)

        payout_cols = [
            ("payout_id", "Payout", lambda r: r["payout_id"]),
            ("arrival_date", "Arrival date", lambda r: r["arrival_date"]),
            ("currency", "Ccy", lambda r: r["currency"]),
            ("charge_count", "# Charges", lambda r: r["charge_count"]),
            ("gross", "Gross", lambda r: _dollars(r["gross_cents"])),
            ("stripe_fee", "Stripe fee", lambda r: _dollars(r["stripe_fee_cents"])),
            ("refunds", "Refunds/adj", lambda r: _dollars(r["refund_net_cents"])),
            ("deposit", "Bank deposit", lambda r: _dollars(r["deposit_cents"])),
            ("events", "Events", lambda r: ", ".join(e["event_name"] for e in r["events"])),
        ]
        ws2 = wb.create_sheet("Payout summary")
        render(ws2, payout_cols, payouts)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def _unix_to_date(value: Any) -> str:
    if not value:
        return ""
    try:
        return datetime.fromtimestamp(int(value), timezone.utc).date().isoformat()
    except (ValueError, OSError, OverflowError, TypeError):
        return ""


def _unix_to_iso(value: Any) -> str:
    if not value:
        return ""
    try:
        return datetime.fromtimestamp(int(value), timezone.utc).isoformat()
    except (ValueError, OSError, OverflowError, TypeError):
        return ""
